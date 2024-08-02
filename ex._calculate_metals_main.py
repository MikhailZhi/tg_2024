import openpyxl

# открываю файл
xlsx_name = r'2024 05 22 p.xlsx'
wb = openpyxl.load_workbook(xlsx_name)

# создается объект с заданного листа книги
xlsx_sheet_name = r'Conc. in Sample Units'
sheet = wb[xlsx_sheet_name]

# получаю номера последнего ряда и последней колонки
row_count = sheet.max_row
column_count = sheet.max_column

# листы для работы переменные
samples = []  # а нужен ли? он же будет в словаре, как ключи
samples_dict = {}
metals_columns = {'As': 'L', 'Cd': 'V', 'Cu': 'AB', 'Ni': 'AS', 'Pb': 'AU', 'Zn': 'BM'}  # Связь металлов и столбцов

print(f"A bond between metals and columns: {metals_columns}")

for i in range(1, row_count + 1):
    sample_code_with_prefix = sheet.cell(i, 2).value.split()[0]  # получение списка из слов в ячейке
    sample_code = sample_code_with_prefix[2:]  # выделяю числовой номер пробы и растворитель
    if sample_code_with_prefix[:2] == "p-":  # перебираю только то, что помечено, как почва
        sample_suffix = sheet.cell(i, 2).value.split()[1]  # получаю суффикс (экстрагент) пробы
        if sample_code not in samples_dict.keys():
            samples_dict[sample_code] = []  # записываю имя пробы и пустой список для значений
        samples_dict[sample_code].append({sample_suffix: {}})  # записываю суффикс

        # Записываю значения по металлам в словарь
        for metal in metals_columns:
            samples_dict[sample_code][len(samples_dict[sample_code]) - 1][sample_suffix][metal] = \
                sheet.cell(i, sheet[metals_columns[metal] + str(i)].column).value

# print('\n', samples_dict, '\n')  # печатаю словарь "как есть", чтобы ориентироваться

# структура:
# {123:[{HCl:{blank:21, Co:34}}, {HF:{blank:32, Co:54}}], 345:[...] } - пример
# { [ {key1: {}, key2: {}}, {}] }
# Красиво печатаю получившийся словарь
for sample_code in samples_dict:
    # print(f'Sample - {sample_code}; ', end='')  # печатаю имя пробы
    for i in range(len(samples_dict[sample_code])):  # перебираю значения в списке, дальше нужен индекс, поэтому так
        for suffix in samples_dict[sample_code][i]:
            print(f'Sample - {sample_code}; Extraction: {suffix}', end='\n\tValues: ')  # печатаю чем экстрагировали
            for met in samples_dict[sample_code][i][suffix]:   # перебираю металлы
                print(f'{met}:{samples_dict[sample_code][i][suffix][met]:.4f} ', end=' ')   # печать значений
            print()

# Сохраняю книгу в файл и закрываю
wb.save(xlsx_name)
wb.close()
