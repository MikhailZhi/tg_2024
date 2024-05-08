import openpyxl

xlsx_name = r'Investment_dataset.xlsx'
wb = openpyxl.load_workbook(xlsx_name)  # открываю файл эксель

xlsx_sheet_investments = r'Current_investments'
sheet = wb[xlsx_sheet_investments]  # загружаю лист текущих инвестиций

# sheet.insert_rows(sheet.max_row+1)
row_count = sheet.max_row
column_titles = ['ID', 'Date', 'Time', 'Title', 'Cost', 'Target']

for i in range(1, row_count+1):
    print(f'\t{i} \t- {sheet.cell(i, 1).value}', end='; ')

for i in range(1, sheet.max_column+1):
    print(f'column {i}, Ok') if sheet.cell(1, i).value == column_titles[i - 1] else print(f'column {i}, NOT Ok')

# lst = []
# for i in range(1, sheet.max_column+1):
#     print(sheet.cell(1, i).value)
#     lst.append(sheet.cell(1, i).value)
# print(lst)

# Не нравится, да и есть вопросы к тому, что сделает
# 1. проверить - что же сделает 2. взять другой способ заполнения
'''for i in range(row_count, row_count + 1):  # добавьте необходимое количество столбцов
    sheet.cell(row=i, column=1).value = 'Значение1'
    sheet.cell(row=i, column=2).value = 'Значение2' '''

# sheet['A' + str(row_count+1)] = row_count * 2

wb.save(xlsx_name)
wb.close()
