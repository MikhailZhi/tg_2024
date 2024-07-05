import copy
import openpyxl

xlsx_name = r'2024 05 22 p.xlsx'
wb = openpyxl.load_workbook(xlsx_name)

xlsx_sheet_name = r'Conc. in Sample Units'
sheet = wb[xlsx_sheet_name]

row_count = sheet.max_row
column_count = sheet.max_column

samples = []
samples_dict_initial = {}
metals_columns = {'As': 'L', 'Cd': 'V', 'Cu': 'AB', 'Ni': 'AS', 'Pb': 'AU', 'Zn': 'BM'}
print(f"A bond between metals and columns: {metals_columns}")

sample_name = sheet.cell(5, 2).value
sample_code_with_prefix = sample_name.split()[0]
print(f'A sample_name: {sample_name}')
print("sample_code_with_suffix for the sample:", sample_code_with_prefix)
print()

# print(sample_code[:1])  # такой срез дает префикс названия пробы
# sheet.cell(i, 2).value  # покажет названия проб

# запись значений из рабочей книги в словарь
for i in range(1, row_count + 1):
    sample_name = sheet.cell(i, 2).value
    # print(f'i={i}, {sheet.cell(i, 2).coordinate} = {sample_name}')
    sample_code_with_prefix = sample_name.split()[0]  # получение списка из слов в ячейке
    sample_code = sample_code_with_prefix[2:]
    if sample_code_with_prefix[:2] == "p-":
        sample_suffix = sheet.cell(i, 2).value.split()[1]  # получаю суффикс пробы
        # samples_dict_initial[sample_code] = {sample_suffix: {}}  # записал суффикс в словарь

        # Записываю значения по металлам в словарь
        for metal in metals_columns:
            samples_dict_initial[sample_code][sample_suffix][metal] = \
                sheet.cell(i, sheet[metals_columns[metal] + str(i)].column).value
            print(f'sample name={sample_code}, sample suffix={sample_suffix}, metal={metal}, '
                  f'value={samples_dict_initial[sample_code][sample_suffix][metal]}')

# просто печать словаря с данными
print(f"A list of samples from the file: {samples}", f"A samples dictionary: {samples_dict_initial}\n", sep='\n')

# красиво печатаю получившийся словарь
print('samples_dict_initial')
for sample in samples_dict_initial:
    print(f'Sample - {sample}; ', end='')
    for suffix in samples_dict_initial[sample]:
        print(f'suffix = {suffix}:')
        for metal in samples_dict_initial[sample][suffix]:
            if samples_dict_initial[sample][suffix][metal] <= 0:
                samples_dict_initial[sample][suffix][metal] = 0
            print(f'metal - {metal}, concentration = {(samples_dict_initial[sample][suffix][metal]):.3f}')
    print()

# Формирую словарь с результатами
samples_dict_recalculated = copy.deepcopy(samples_dict_initial)
del samples_dict_recalculated['blank']

# вычитаю бланки, переписываю значения
for sample in samples_dict_recalculated:
    for suffix in samples_dict_recalculated[sample]:
        for metal in samples_dict_recalculated[sample][suffix]:
            samples_dict_recalculated[sample][suffix][metal] = (
                    samples_dict_initial[sample][suffix][metal] - samples_dict_initial['blank'][suffix][metal])

# красиво печатаю пересчитанный словарь
for sample in samples_dict_recalculated:
    print(f'Sample - {sample}; ', end='')
    for suffix in samples_dict_recalculated[sample]:
        print(f'suffix = {suffix}:')
        for metal in samples_dict_recalculated[sample][suffix]:
            print(f'metal - {metal}, concentration = {(samples_dict_recalculated[sample][suffix][metal]):.3f},'
                  f' было - {(samples_dict_initial[sample][suffix][metal]):.3f}')
    print()


# print(sheet.cell(1, 20).coordinate, sheet['AU1'].column)  # показывает координаты и столбец ячейки

# тренируюсь грузить данные в словарь
# sample_name = sheet.cell(5, 2).value
# sample_cell_name = sample_name.split()
# sample_suffix = sample_cell_name[1]
# print(f'A suffix for a sample: {sample_suffix}')
# samples_dict_2 = samples_dict
# print(f'\nA sample dict before: {samples_dict_2}')
# samples_dict_2[sample_cell_name[0][2:]] = {sample_suffix: {}}
# print(f'A sample dict after: {samples_dict_2}')

# пробую грузить в словарь


# проверка правильных названий столбцов

# sheet.cell(1, i).value  # покажет названия столбцов
# for i in range(1, column_count + 1):
#     print(f'i={i}, column={sheet.cell(1, i).value}')

wb.save(xlsx_name)
wb.close()
