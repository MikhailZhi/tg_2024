from aiogram import Bot, Dispatcher
from aiogram.filters import Command
from aiogram.fsm.state import StatesGroup, State
import openpyxl
# from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton

BOT_TOKEN = '5785150011:AAFfBxg0EpqYDtYcuARILwXY8BDlk-_qQzs'  # https://t.me/Mike_sdbot reloc_sd_bot

# Создаем объекты бота и диспетчера
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


# Создаю класс конечных состояний для приема и обработки инвестиций
class Investment(StatesGroup):
    new_investment_title = State()
    new_investment_cost = State()
    new_investment_target = State()


# список для хранения полученных данных о новой инвестиции
new_investment_data = []


@dp.message(Command(commands='start'))  # Обработчик команды start
async def command_start(message):
    print('start')
    # await bot.send_message(chat_id=message.chat.id, text='Hello!', reply_to_message_id=message.message_id)
    await message.reply(text='''Я знаю команды:
                             /start
                             /new_investment
                             /cancel''')


@dp.message(Command(commands='new_investment'))  # команда для начала ввода новых инвестиций
async def command_new_investment(message, state):
    await message.reply(text='Введите название инвестиции')
    await state.set_state(Investment.new_investment_title)


@dp.message(Command(commands='cancel'))  # команда для отмены текущего состояния
async def command_cancel(message, state):
    await message.reply(text='Действие отменено.')

    await state.clear()


# handler
@dp.message(Investment.new_investment_title)
async def new_investment_title(message, state):
    new_investment_data.append(message.text.lower())
    print(new_investment_data)
    await message.reply(text=f'Для инвестиции "{new_investment_data[0]}"\n'
                             f'Введите цену вложения'
                             f'PS - целое число, без пробелов, в рублях')
    await state.set_state(Investment.new_investment_cost)


# handler
@dp.message(Investment.new_investment_cost)
async def new_investment_cost(message, state):
    new_investment_data.append(message.text)
    print(new_investment_data)
    await message.reply(text='Введите ожидаемую цену продажи:'
                             'PS - целое число, без пробелов, в рублях')
    await state.set_state(Investment.new_investment_target)


# handler
@dp.message(Investment.new_investment_target)
async def new_investment_target(message, state):
    new_investment_data.append(message.text)
    print(new_investment_data)
    await message.answer(text=f'Для инвестиции {new_investment_data[0]}\n'
                              f' Вложения составляют {new_investment_data[1]}\n'
                              f' Цена продажи ожидается {new_investment_data[2]}')
    await xls()
    await state.clear()

# переменные для работы с таблицей
xlsx_name = r'Investment_dataset.xlsx'
xlsx_sheet_investments = r'Current_investments'
xlsx_column_titles = ['ID', 'Date', 'Time', 'Title', 'Cost', 'Target']


# функция для записи новой инвестиции в файл эксель
async def xls():
    wb = openpyxl.load_workbook(xlsx_name)  # открываю файл эксель
    sheet = wb[xlsx_sheet_investments]  # загружаю лист текущих инвестиций
    row_count = sheet.max_row

    for i in range(1, row_count + 1):
        print(f'\t{i} \t- {sheet.cell(i, 1).value}', end='; ')  # печать содержимого колонки 1
    print()

    for i in range(1, sheet.max_column + 1):  # проверка, что колонки названы верно
        print(f'column {i}, Ok') if sheet.cell(1, i).value == xlsx_column_titles[i - 1] \
            else print(f'column {i}, NOT Ok')

    wb.save(xlsx_name)
    wb.close()

if __name__ == '__main__':
    print('Запускаем!..')
    dp.run_polling(bot)
